import argparse
import json
import os
import re
import sys
import textwrap

from bs4 import BeautifulSoup
from openpyxl import Workbook
import requests


def get_soup_from_url(url):
    request = requests.get(url)
    request.encoding = 'utf-8'
    if request.status_code == 200:
        return BeautifulSoup(request.text, 'lxml')


def get_courses_list(xml_feed_url, amount_of_courses):
    courses_list_soup = get_soup_from_url(xml_feed_url)
    if courses_list_soup:
        courses_urls = courses_list_soup.find_all('loc',
                                                  limit=amount_of_courses)
        list_of_courses_urls = [tag.text for tag in courses_urls]
        return list_of_courses_urls


def parse_courses(urls_list):
    list_of_dictionaries_with_courses_info = []
    for url in urls_list:
        course_soup = get_soup_from_url(url)
        course_info = get_course_info(course_soup)
        list_of_dictionaries_with_courses_info.append(course_info)
    return list_of_dictionaries_with_courses_info


def get_course_name(soup):
    course_name_class = 'title display-3-text'
    try:
        return soup.find(class_=course_name_class).text
    except AttributeError:
        return


def get_course_language(soup):
    course_language_class = 'language-info'
    try:
        return soup.find(class_=course_language_class).text
    except AttributeError:
        return


def get_course_duration(soup):
    course_weeks_tag = 'rc-WeekView'
    week_tag = 'week'
    try:
        weeks_list = soup.find(class_=course_weeks_tag).find_all(
            class_=week_tag)
        return len(weeks_list)
    except AttributeError:
        return


def get_course_average_score(soup):
    try:
        script_tag = soup.find('script', text=re.compile('window.App')).text
        return re.findall(r'"averageFiveStarRating":([\d.]+)',
                          script_tag)[0]
    except (AttributeError, IndexError):
        return


def get_course_start_date(soup):
    try:
        course_script_vars = json.loads(soup.find('script',
                                        type='application/ld+json').text.
                                        strip())
        return course_script_vars['hasCourseInstance'][0]['startDate']
    except (AttributeError, IndexError, KeyError):
        return


def get_course_info(soup):
    course_dictionary = {'name': get_course_name(soup),
                         'language': get_course_language(soup),
                         'weeks': get_course_duration(soup),
                         'average_score': get_course_average_score(soup),
                         'start_date': get_course_start_date(soup)}
    return course_dictionary


def output_courses_info_to_xlsx(workbook, save_file_path, courses_info,
                                name_column_width, language_column_width):
    sheet = workbook.active
    for row, course in enumerate(courses_info):
        sheet.cell(row=row + 2, column=1).value = textwrap.fill(
            course['name'], name_column_width)
        sheet.cell(row=row + 2, column=2).value = textwrap.fill(
            course['language'], language_column_width)
        sheet.cell(row=row + 2, column=3).value = course['start_date']
        sheet.cell(row=row + 2, column=4).value = course['weeks']
        sheet.cell(row=row + 2, column=5).value = course['average_score']
    save_file_path = os.path.join(save_file_path)
    workbook.save(save_file_path)


def setup_excel_workbook(name_width, language_width, date_width=16,
                         duration_width=9, average_score_width=16):
    workbook = Workbook()
    sheet = workbook.active
    sheet.column_dimensions['A'].width = name_width
    sheet.column_dimensions['B'].width = language_width
    sheet.column_dimensions['C'].width = date_width
    sheet.column_dimensions['D'].width = duration_width
    sheet.column_dimensions['E'].width = average_score_width
    sheet.cell(row=1, column=1).value = 'Name of course'
    sheet.cell(row=1, column=2).value = 'Language'
    sheet.cell(row=1, column=3).value = 'Date of start'
    sheet.cell(row=1, column=4).value = 'Duration'
    sheet.cell(row=1, column=5).value = 'Average score'
    return workbook


def parse_arguments():
    parser = argparse.ArgumentParser(description='Write Coursera courses info'
                                                 ' to xlsx')
    parser.add_argument('amount', nargs='?', default=20,
                        type=int, help='amount of courses to parse')
    parser.add_argument('output', nargs='?', default='courses.xlsx', type=str,
                        help='name of output file')
    parser.add_argument('name_width', nargs='?', default=40, type=int,
                        help='width of name column in spreadsheet')
    parser.add_argument('lang_width', nargs='?', default=40, type=int,
                        help='width of language column in spreadsheet')
    arguments = parser.parse_args()
    return arguments.amount, arguments.name_width, arguments.lang_width,\
        arguments.output


def append_xlsx_file_extension_if_needed(file_name):
    return file_name + '.xlsx' if '.xlsx' not in file_name else file_name


if __name__ == '__main__':
    coursera_xml_feed_url = 'https://www.coursera.org/sitemap~www~courses.xml'
    courses_amount, name_col_width, language_col_width, output_file_name =\
        parse_arguments()

    output_file_name = append_xlsx_file_extension_if_needed(output_file_name)

    print('---Getting course list...')
    courses_urls_list = get_courses_list(coursera_xml_feed_url, courses_amount)

    sys.exit('!!! Something went wrong: check your internet connection;'
             ' also coursera xml feed url may changed'
             ) if not courses_urls_list else None

    courses = parse_courses(courses_urls_list)
    print('---Parsing courses info. This can take awhile (few seconds for'
          ' every course).')

    wb = setup_excel_workbook(name_col_width, language_col_width)
    output_courses_info_to_xlsx(wb, output_file_name, courses,
                                name_col_width, language_col_width)
    print('---Success: {} is ready!'.format(output_file_name))
